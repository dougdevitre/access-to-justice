"""
Companion budget workbook for Doc 13 — Budget Narrative & Line-Items.
Produces /mnt/user-data/outputs/13b-budget-workbook.xlsx with 4 sheets:
  - Summary        (category × year, formula-driven totals)
  - Detail         (line items with role rates + FTE, formula-driven)
  - Funding Sources (funder stack with 20% cap check)
  - Assumptions    (text-documented assumptions with sources)

All totals are Excel formulas, not hardcoded. Inputs are blue-text.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- Style constants ----------
INK       = '0D1220'
ACCENT    = 'B8371F'
BLUE      = '0000FF'   # industry-standard "hardcoded input"
GREEN     = '008000'   # industry-standard "link from another sheet"
PAPER_SOFT = 'F5EFE5'
RULE      = 'CCCCCC'

thin  = Side(border_style='thin', color=RULE)
heavy = Side(border_style='medium', color=INK)
box   = Border(left=thin, right=thin, top=thin, bottom=thin)

FONT_FAMILY = 'Arial'

def header_style(cell):
    cell.font = Font(name=FONT_FAMILY, bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill('solid', start_color=INK)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = box

def label_style(cell, bold=True):
    cell.font = Font(name=FONT_FAMILY, bold=bold, color=INK, size=11)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = box

def input_style(cell):
    cell.font = Font(name=FONT_FAMILY, color=BLUE, size=11)
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.number_format = '$#,##0;($#,##0);-'
    cell.border = box

def formula_style(cell):
    cell.font = Font(name=FONT_FAMILY, color=INK, size=11)
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.number_format = '$#,##0;($#,##0);-'
    cell.border = box

def link_style(cell):
    cell.font = Font(name=FONT_FAMILY, color=GREEN, size=11)
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.number_format = '$#,##0;($#,##0);-'
    cell.border = box

def total_style(cell):
    cell.font = Font(name=FONT_FAMILY, bold=True, color=INK, size=11)
    cell.fill = PatternFill('solid', start_color=PAPER_SOFT)
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.number_format = '$#,##0;($#,##0);-'
    cell.border = Border(left=thin, right=thin, top=heavy, bottom=heavy)

def pct_style(cell):
    cell.font = Font(name=FONT_FAMILY, color=INK, size=11)
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.number_format = '0.0%'
    cell.border = box

def title_style(cell):
    cell.font = Font(name=FONT_FAMILY, bold=True, color=INK, size=16)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def subtitle_style(cell):
    cell.font = Font(name=FONT_FAMILY, color=ACCENT, size=10, bold=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')

wb = Workbook()

# =====================================================
# Sheet 1: Detail — line items feed the Summary
# =====================================================
det = wb.active
det.title = 'Detail'

det.column_dimensions['A'].width = 38
det.column_dimensions['B'].width = 10
det.column_dimensions['C'].width = 14
det.column_dimensions['D'].width = 14
det.column_dimensions['E'].width = 14
det.column_dimensions['F'].width = 14
det.column_dimensions['G'].width = 38

title_style(det['A1']); det['A1'] = 'Family Court Metric Catalog — Missouri Pilot'
subtitle_style(det['A2']); det['A2'] = 'DETAIL · LINE ITEMS (3-YEAR BUDGET)'

# Header row at row 4
headers = ['Line Item', 'FTE / Units', 'Year 1', 'Year 2', 'Year 3', 'Total (3yr)', 'Source / Notes']
for i, h in enumerate(headers, start=1):
    c = det.cell(row=4, column=i, value=h)
    header_style(c)

# Line items - (label, category, fte/units display, y1, y2, y3, notes)
# Categories: PERSONNEL, BOARD, INFRA, ACADEMIC, CONTINGENCY
lines = [
    # label, category, fte, y1, y2, y3, note
    ('Executive Director',                      'PERSONNEL', '1.0 FTE', 130000, 135000, 141000, 'Nonprofit ED, STL market rate'),
    ('Engineering Lead (Founder)',              'PERSONNEL', '1.0 FTE', 95000,  99000,  103000, 'Below-market founder salary'),
    ('Methodology Coordinator',                 'PERSONNEL', '0.5 FTE', 43000,  45000,  47000,  'Part-time research-ops role'),
    ('Research / Data Analyst',                 'PERSONNEL', '0.5 FTE', 35000,  37000,  39000,  'Part-time analytics role'),
    ('Contract — Privacy Audit',                'PERSONNEL', '0.2 FTE', 15000,  15000,  15000,  'External auditor, annual engagement'),
    ('Methodology Board Honoraria',             'BOARD',     '11 × 4',  22000,  22000,  22000,  '11 members × 4 meetings × $500'),
    ('Data Ethics Review Honoraria',            'BOARD',     '5 × 2',   20000,  22000,  24000,  '5 members × ~2 reviews × $2,000'),
    ('AWS Infrastructure',                      'INFRA',     '—',       15000,  16200,  17500,  'DynamoDB, Lambda, API Gtwy, S3, SSM'),
    ('Third-Party Tools (observability, etc.)', 'INFRA',     '—',       8500,   9000,   9500,   'Datadog/Sentry/compliance tooling'),
    ('Independent Security Audit',              'INFRA',     'Annual',  12000,  12000,  12000,  'Required by MOU before Month 6'),
    ('Insurance (Cyber $2M, E&O)',              'INFRA',     '—',       8000,   8000,   8000,   'Required by MOU Article IX'),
    ('Office / Meeting Costs',                  'INFRA',     '—',       3000,   3000,   3000,   'Quarterly Board meeting logistics'),
    ('Wash U Law Research Contract',            'ACADEMIC',  'Annual',  15000,  15000,  15000,  'Grad research + IRB support'),
    ('Mizzou Law Research Contract',            'ACADEMIC',  'Annual',  15000,  15000,  15000,  'Grad research + IRB support'),
    ('Missouri State CPA Research Contract',    'ACADEMIC',  'Annual',  15000,  15000,  15000,  'Grad research + IRB support'),
    ('Contingency (10% of operating)',          'CONTINGENCY','—',      15810,  14970,  18030,  'Unallocated reserve, Board-released'),
]

row = 5
first_line_row = row
personnel_rows = []
board_rows = []
infra_rows = []
academic_rows = []
contingency_rows = []

for (label, category, fte, y1, y2, y3, note) in lines:
    det.cell(row=row, column=1, value=label); label_style(det.cell(row=row, column=1), bold=False)
    det.cell(row=row, column=2, value=fte)
    c = det.cell(row=row, column=2); c.font = Font(name=FONT_FAMILY, color=INK, size=11); c.alignment = Alignment(horizontal='center'); c.border = box

    for col, val in [(3, y1), (4, y2), (5, y3)]:
        cc = det.cell(row=row, column=col, value=val)
        input_style(cc)

    # Total (3yr) — formula
    tc = det.cell(row=row, column=6, value=f'=SUM(C{row}:E{row})')
    formula_style(tc)

    # Notes
    nc = det.cell(row=row, column=7, value=note)
    nc.font = Font(name=FONT_FAMILY, color='555046', size=10, italic=False)
    nc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    nc.border = box

    if category == 'PERSONNEL': personnel_rows.append(row)
    elif category == 'BOARD': board_rows.append(row)
    elif category == 'INFRA': infra_rows.append(row)
    elif category == 'ACADEMIC': academic_rows.append(row)
    elif category == 'CONTINGENCY': contingency_rows.append(row)

    row += 1

last_line_row = row - 1

# Subtotal rows by category
row += 1
subtotal_row_start = row

def write_subtotal(label, row_list):
    global row
    det.cell(row=row, column=1, value=f'Subtotal — {label}').font = Font(name=FONT_FAMILY, bold=True, color=INK, size=11)
    det.cell(row=row, column=1).alignment = Alignment(horizontal='left')
    det.cell(row=row, column=1).border = box
    # Build SUM formula from row list
    if row_list:
        cells_y1 = ','.join([f'C{r}' for r in row_list])
        cells_y2 = ','.join([f'D{r}' for r in row_list])
        cells_y3 = ','.join([f'E{r}' for r in row_list])
        det.cell(row=row, column=3, value=f'=SUM({cells_y1})')
        det.cell(row=row, column=4, value=f'=SUM({cells_y2})')
        det.cell(row=row, column=5, value=f'=SUM({cells_y3})')
        det.cell(row=row, column=6, value=f'=SUM(C{row}:E{row})')
        for col in [3, 4, 5, 6]:
            c = det.cell(row=row, column=col)
            c.font = Font(name=FONT_FAMILY, bold=True, color=INK, size=11)
            c.fill = PatternFill('solid', start_color=PAPER_SOFT)
            c.alignment = Alignment(horizontal='right')
            c.number_format = '$#,##0;($#,##0);-'
            c.border = Border(left=thin, right=thin, top=heavy, bottom=thin)
    row += 1
    return row - 1

pers_sub = write_subtotal('Personnel', personnel_rows)
board_sub = write_subtotal('Board & Reviews', board_rows)
infra_sub = write_subtotal('Infrastructure & Ops', infra_rows)
acad_sub = write_subtotal('Academic Partnerships', academic_rows)
cont_sub = write_subtotal('Contingency', contingency_rows)

# Grand Total
row += 1
grand_row = row
det.cell(row=row, column=1, value='GRAND TOTAL (3-YEAR PILOT)')
det.cell(row=row, column=1).font = Font(name=FONT_FAMILY, bold=True, color='FFFFFF', size=12)
det.cell(row=row, column=1).fill = PatternFill('solid', start_color=INK)
det.cell(row=row, column=1).alignment = Alignment(horizontal='left')
det.cell(row=row, column=1).border = box

for col in [3, 4, 5]:
    cell_letter = get_column_letter(col)
    # Sum of subtotal rows
    f = f'={cell_letter}{pers_sub}+{cell_letter}{board_sub}+{cell_letter}{infra_sub}+{cell_letter}{acad_sub}+{cell_letter}{cont_sub}'
    det.cell(row=row, column=col, value=f)
    c = det.cell(row=row, column=col)
    c.font = Font(name=FONT_FAMILY, bold=True, color='FFFFFF', size=12)
    c.fill = PatternFill('solid', start_color=INK)
    c.alignment = Alignment(horizontal='right')
    c.number_format = '$#,##0;($#,##0);-'
    c.border = box

det.cell(row=row, column=6, value=f'=SUM(C{row}:E{row})')
c = det.cell(row=row, column=6)
c.font = Font(name=FONT_FAMILY, bold=True, color='FFFFFF', size=12)
c.fill = PatternFill('solid', start_color=INK)
c.alignment = Alignment(horizontal='right')
c.number_format = '$#,##0;($#,##0);-'
c.border = box

# Freeze panes
det.freeze_panes = 'C5'

# =====================================================
# Sheet 2: Summary — category × year, pulls from Detail via cross-sheet links
# =====================================================
summary = wb.create_sheet('Summary', 0)

summary.column_dimensions['A'].width = 32
for col in ['B', 'C', 'D', 'E']:
    summary.column_dimensions[col].width = 15
summary.column_dimensions['F'].width = 12

title_style(summary['A1']); summary['A1'] = 'Family Court Metric Catalog — Missouri Pilot'
subtitle_style(summary['A2']); summary['A2'] = 'SUMMARY · 3-YEAR BUDGET BY CATEGORY'

# Headers at row 4
headers = ['Category', 'Year 1', 'Year 2', 'Year 3', 'Total (3yr)', '% of Total']
for i, h in enumerate(headers, start=1):
    c = summary.cell(row=4, column=i, value=h)
    header_style(c)

# Categories pull from Detail subtotals
summary_cats = [
    ('Personnel',               pers_sub),
    ('Board & Reviews',         board_sub),
    ('Infrastructure & Ops',    infra_sub),
    ('Academic Partnerships',   acad_sub),
    ('Contingency',             cont_sub),
]

row = 5
for (cat, detail_row) in summary_cats:
    summary.cell(row=row, column=1, value=cat)
    label_style(summary.cell(row=row, column=1), bold=False)
    # Green links to Detail sheet
    summary.cell(row=row, column=2, value=f"=Detail!C{detail_row}"); link_style(summary.cell(row=row, column=2))
    summary.cell(row=row, column=3, value=f"=Detail!D{detail_row}"); link_style(summary.cell(row=row, column=3))
    summary.cell(row=row, column=4, value=f"=Detail!E{detail_row}"); link_style(summary.cell(row=row, column=4))
    # Row total
    summary.cell(row=row, column=5, value=f'=SUM(B{row}:D{row})'); formula_style(summary.cell(row=row, column=5))
    # % of total — filled after grand total row is placed
    row += 1

# Totals row
total_row = row + 1
summary.cell(row=total_row, column=1, value='TOTAL')
c = summary.cell(row=total_row, column=1); c.font = Font(name=FONT_FAMILY, bold=True, color='FFFFFF', size=12); c.fill = PatternFill('solid', start_color=INK); c.alignment = Alignment(horizontal='left'); c.border = box

for col in [2, 3, 4, 5]:
    cell_letter = get_column_letter(col)
    summary.cell(row=total_row, column=col, value=f'=SUM({cell_letter}5:{cell_letter}{row-1})')
    cc = summary.cell(row=total_row, column=col)
    cc.font = Font(name=FONT_FAMILY, bold=True, color='FFFFFF', size=12)
    cc.fill = PatternFill('solid', start_color=INK)
    cc.alignment = Alignment(horizontal='right')
    cc.number_format = '$#,##0;($#,##0);-'
    cc.border = box

# % of total column for each category row
for r in range(5, 5 + len(summary_cats)):
    summary.cell(row=r, column=6, value=f'=E{r}/$E${total_row}')
    pct_style(summary.cell(row=r, column=6))

# % total row cell
summary.cell(row=total_row, column=6, value=f'=E{total_row}/E{total_row}')
cc = summary.cell(row=total_row, column=6); cc.font = Font(name=FONT_FAMILY, bold=True, color='FFFFFF', size=12); cc.fill = PatternFill('solid', start_color=INK); cc.alignment = Alignment(horizontal='right'); cc.number_format = '0.0%'; cc.border = box

# Narrative note
note_row = total_row + 3
summary.cell(row=note_row, column=1, value='Budget Integrity')
c = summary.cell(row=note_row, column=1); c.font = Font(name=FONT_FAMILY, bold=True, color=ACCENT, size=11); c.alignment = Alignment(horizontal='left')
summary.cell(row=note_row + 1, column=1, value='All Year 1/2/3 amounts pull directly from Detail sheet. Change a rate or FTE in Detail; all totals update automatically.')
summary.cell(row=note_row + 1, column=1).font = Font(name=FONT_FAMILY, color=INK, size=10, italic=True)
summary.cell(row=note_row + 1, column=1).alignment = Alignment(horizontal='left', wrap_text=True)

summary.merge_cells(start_row=note_row + 1, start_column=1, end_row=note_row + 1, end_column=6)

# =====================================================
# Sheet 3: Funding Sources — with 20% cap check
# =====================================================
funding = wb.create_sheet('Funding Sources')

funding.column_dimensions['A'].width = 38
funding.column_dimensions['B'].width = 12
funding.column_dimensions['C'].width = 16
funding.column_dimensions['D'].width = 16
funding.column_dimensions['E'].width = 30

title_style(funding['A1']); funding['A1'] = 'Funding Sources · 20% Cap Enforced'
subtitle_style(funding['A2']); funding['A2'] = 'NO SINGLE FUNDER MAY EXCEED 20% OF TOTAL BUDGET'

# Pull grand total from Summary
funding.cell(row=4, column=1, value='Total Budget (from Summary)')
label_style(funding.cell(row=4, column=1), bold=True)
funding.cell(row=4, column=2, value=f'=Summary!E{total_row}')
c = funding.cell(row=4, column=2); c.font = Font(name=FONT_FAMILY, color=GREEN, bold=True, size=11); c.number_format = '$#,##0'; c.alignment = Alignment(horizontal='right'); c.border = box

funding.cell(row=5, column=1, value='20% Cap Threshold')
label_style(funding.cell(row=5, column=1), bold=True)
funding.cell(row=5, column=2, value='=B4*0.2')
c = funding.cell(row=5, column=2); c.font = Font(name=FONT_FAMILY, color=INK, bold=True, size=11); c.number_format = '$#,##0'; c.alignment = Alignment(horizontal='right'); c.border = box; c.fill = PatternFill('solid', start_color='FFF3CD')

# Funders table header
for i, h in enumerate(['Funder Category', 'Target %', 'Target $', 'Secured $', 'Status / Notes'], start=1):
    cc = funding.cell(row=7, column=i, value=h); header_style(cc)

funders = [
    ('Philanthropy — Access-to-Justice Foundations', 0.40, 'Active cultivation · multi-funder stack'),
    ('Academic Institutional Partnerships',          0.15, 'Contracted at partner level (Wash U, Mizzou, MO State)'),
    ('Individual & Community Donors',                0.15, 'Ongoing direct fundraising'),
    ('CoTrackPro Subscription Revenue',              0.10, 'Existing modest revenue stream'),
    ('Deferred — State Appropriation (post-pilot)',  0.20, 'Not Phase 1; placeholder pending legislative action'),
]

row = 8
for (name, pct, note) in funders:
    funding.cell(row=row, column=1, value=name); label_style(funding.cell(row=row, column=1), bold=False)
    funding.cell(row=row, column=2, value=pct); c = funding.cell(row=row, column=2); c.number_format = '0.0%'; c.font = Font(name=FONT_FAMILY, color=BLUE, size=11); c.alignment = Alignment(horizontal='right'); c.border = box
    funding.cell(row=row, column=3, value=f'=B{row}*$B$4'); formula_style(funding.cell(row=row, column=3))
    funding.cell(row=row, column=4, value=0); input_style(funding.cell(row=row, column=4))
    funding.cell(row=row, column=5, value=note); c = funding.cell(row=row, column=5); c.font = Font(name=FONT_FAMILY, color='555046', size=10); c.alignment = Alignment(horizontal='left', wrap_text=True); c.border = box
    row += 1

# Totals
total_funding_row = row + 1
funding.cell(row=total_funding_row, column=1, value='TOTAL')
c = funding.cell(row=total_funding_row, column=1); c.font = Font(name=FONT_FAMILY, bold=True, color='FFFFFF', size=12); c.fill = PatternFill('solid', start_color=INK); c.border = box
funding.cell(row=total_funding_row, column=2, value=f'=SUM(B8:B{row-1})')
c = funding.cell(row=total_funding_row, column=2); c.font = Font(name=FONT_FAMILY, bold=True, color='FFFFFF', size=12); c.fill = PatternFill('solid', start_color=INK); c.number_format = '0.0%'; c.alignment = Alignment(horizontal='right'); c.border = box
funding.cell(row=total_funding_row, column=3, value=f'=SUM(C8:C{row-1})')
c = funding.cell(row=total_funding_row, column=3); c.font = Font(name=FONT_FAMILY, bold=True, color='FFFFFF', size=12); c.fill = PatternFill('solid', start_color=INK); c.number_format = '$#,##0'; c.alignment = Alignment(horizontal='right'); c.border = box
funding.cell(row=total_funding_row, column=4, value=f'=SUM(D8:D{row-1})')
c = funding.cell(row=total_funding_row, column=4); c.font = Font(name=FONT_FAMILY, bold=True, color='FFFFFF', size=12); c.fill = PatternFill('solid', start_color=INK); c.number_format = '$#,##0'; c.alignment = Alignment(horizontal='right'); c.border = box

# Cap-check formula rows
check_start = total_funding_row + 3
funding.cell(row=check_start, column=1, value='Cap Check — Any Funder Over 20%?')
c = funding.cell(row=check_start, column=1); c.font = Font(name=FONT_FAMILY, bold=True, color=ACCENT, size=11); c.alignment = Alignment(horizontal='left')

funding.cell(row=check_start + 1, column=1, value='Max single funder target %')
funding.cell(row=check_start + 1, column=2, value=f'=MAX(B8:B{row-1})')
c = funding.cell(row=check_start + 1, column=2); c.font = Font(name=FONT_FAMILY, bold=True, color=INK, size=11); c.number_format = '0.0%'; c.alignment = Alignment(horizontal='right'); c.border = box

funding.cell(row=check_start + 2, column=1, value='Cap violation?')
funding.cell(row=check_start + 2, column=2, value=f'=IF(B{check_start+1}>0.2,"YES — REBALANCE","OK")')
c = funding.cell(row=check_start + 2, column=2); c.font = Font(name=FONT_FAMILY, bold=True, color=INK, size=11); c.alignment = Alignment(horizontal='right'); c.border = box; c.fill = PatternFill('solid', start_color='D4EDDA')

# =====================================================
# Sheet 4: Assumptions
# =====================================================
asm = wb.create_sheet('Assumptions')
asm.column_dimensions['A'].width = 4
asm.column_dimensions['B'].width = 40
asm.column_dimensions['C'].width = 50
asm.column_dimensions['D'].width = 30

title_style(asm['A1']); asm['A1'] = 'Budget Assumptions'
subtitle_style(asm['A2']); asm['A2'] = 'DOCUMENT ANY CHANGE; SOURCE EVERY ASSUMPTION'

for i, h in enumerate(['#', 'Assumption', 'Rationale', 'Source / Reference'], start=1):
    cc = asm.cell(row=4, column=i, value=h); header_style(cc)

assumptions = [
    (1, 'One executive director hired within 90 days of funding secured', 'Operational leadership is the critical path for Phase 1', 'Standard nonprofit startup timeline'),
    (2, 'No staff growth beyond Year 1 headcount', 'Keeps budget predictable; absorbs growth via contractors only', 'Internal planning decision'),
    (3, 'Pilot circuits require no reimbursement', 'MOU structure assumes voluntary participation', 'OSCA Data-Sharing MOU (Doc 06), Article VIII'),
    (4, 'AWS costs stay within Activate credits for first 6 months', 'New nonprofit qualifies for AWS Activate $5K tier', 'AWS Activate program eligibility'),
    (5, 'Academic partners provide in-kind contribution (faculty time, IRB)', 'Contracted $15k/yr supplements ~3-5x that value in-kind', 'Standard academic partnership pattern'),
    (6, 'No litigation expenses in baseline budget', 'Test cases would require separate supplemental funding', 'Risk Register (Doc 10), Category IV'),
    (7, 'Year 2 and Year 3 include 3% escalation over Year 1', 'Modest inflation adjustment consistent with nonprofit norms', 'BLS CPI projections, conservative'),
    (8, 'Contingency (10%) held in reserve', 'Board-released, not presumed available for operations', 'Internal governance decision'),
    (9, '20% funder cap enforced regardless of category', 'Prevents capture by any single interest', 'Methodology Board Charter (Doc 07)'),
    (10, 'All figures in 2026 US dollars', 'No inflation adjustment beyond stated 3% Y-over-Y', 'Standard nonprofit budget convention'),
]

row = 5
for (num, assumption, rationale, source) in assumptions:
    asm.cell(row=row, column=1, value=num)
    c = asm.cell(row=row, column=1); c.font = Font(name=FONT_FAMILY, bold=True, color=ACCENT, size=11); c.alignment = Alignment(horizontal='center', vertical='top'); c.border = box

    asm.cell(row=row, column=2, value=assumption)
    c = asm.cell(row=row, column=2); c.font = Font(name=FONT_FAMILY, bold=True, color=INK, size=10); c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True); c.border = box

    asm.cell(row=row, column=3, value=rationale)
    c = asm.cell(row=row, column=3); c.font = Font(name=FONT_FAMILY, color=INK, size=10); c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True); c.border = box

    asm.cell(row=row, column=4, value=source)
    c = asm.cell(row=row, column=4); c.font = Font(name=FONT_FAMILY, color='555046', size=9, italic=True); c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True); c.border = box

    asm.row_dimensions[row].height = 36
    row += 1

# Set Summary as active starting sheet
wb.active = 0

# Save
output_path = '/mnt/user-data/outputs/13b-budget-workbook.xlsx'
wb.save(output_path)
print(f'Saved: {output_path}')
